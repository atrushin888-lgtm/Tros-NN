"""
Юнит-экономика Трос NN
Логика:
  Полная себ-ть 1 шт = себ-ть_троса + ручка + доля_доставки_позиции / кол-во
  Цена = полная_себ-ть × 1.4
  Прибыль 1 шт = цена - полная_себ-ть
  Чистая прибыль партии = выручка_всего - налог(6%) - маркетинг - себ-ть_партии
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── ИСХОДНЫЕ ДАННЫЕ ────────────────────────────────────────────────────────

COST_PER_M = {6: 180, 8: 250, 10: 272, 12: 322, 14: 470, 16: 539}
WEIGHT_PER_M = {6: 0.187, 8: 0.326, 10: 0.485, 12: 0.700, 14: 1.05, 16: 1.241}
HANDLE       = 472
DELIVERY     = 5_000
MARKETING    = 10_000
TAX_RATE     = 0.06
MARKUP       = 1.40

PURCHASES = [
    (8,  5,  8),
    (8,  10, 6),
    (6,  5,  4),
    (6,  10, 2),
    (10, 5,  3),
    (10, 10, 3),
    (10, 15, 2),
    (12, 10, 3),
    (12, 15, 2),
    (12, 20, 2),
    (14, 15, 2),
    (14, 20, 1),
    (16, 10, 1),
    (16, 15, 1),
]

# ── РАСЧЁТ ─────────────────────────────────────────────────────────────────

rows = []
for diam, length, qty in PURCHASES:
    rope_cost      = COST_PER_M[diam] * length          # себ-ть троса 1 шт
    wpm            = WEIGHT_PER_M[diam]
    weight_pos     = wpm * length * qty                  # вес всей позиции
    rows.append({
        "diam":       diam,
        "length":     length,
        "qty":        qty,
        "cpm":        COST_PER_M[diam],
        "rope_cost":  rope_cost,
        "handle":     HANDLE,
        "wpm":        wpm,
        "weight_pos": weight_pos,
    })

total_weight = sum(r["weight_pos"] for r in rows)

for r in rows:
    deliv_pos       = DELIVERY * r["weight_pos"] / total_weight
    deliv_per_unit  = deliv_pos / r["qty"]
    full_cost       = r["rope_cost"] + r["handle"] + deliv_per_unit
    price           = full_cost * MARKUP
    profit_unit     = price - full_cost
    profit_pos      = profit_unit * r["qty"]

    r["deliv_pos"]      = deliv_pos
    r["deliv_per_unit"] = deliv_per_unit
    r["full_cost"]      = full_cost
    r["price"]          = price
    r["profit_unit"]    = profit_unit
    r["profit_pos"]     = profit_pos

# Итоги партии
total_qty       = sum(r["qty"]        for r in rows)
total_rope_cost = sum(r["rope_cost"] * r["qty"] for r in rows)
handle_total    = HANDLE * total_qty
total_batch_cost= sum(r["full_cost"] * r["qty"] for r in rows)   # включает доставку
revenue         = sum(r["price"]     * r["qty"] for r in rows)
tax             = revenue * TAX_RATE
net_profit      = revenue - tax - MARKETING - total_batch_cost

margin_pct      = net_profit / revenue * 100 if revenue else 0

sorted_by_profit = sorted(rows, key=lambda r: r["profit_pos"], reverse=True)

# ── СТИЛИ ──────────────────────────────────────────────────────────────────

C_DARK    = "1F3864"
C_BLUE    = "2E75B6"
C_ORANGE  = "C55A11"
C_GREEN   = "375623"
C_RED     = "9C0006"
C_ALT     = "EBF3FB"
C_WHITE   = "FFFFFF"
C_YELLOW  = "FFF2CC"
C_GFILL   = "E2EFDA"
C_RFILL   = "FCE4D6"
C_HEADER  = "D6E4F7"

RUB  = '#,##0 ₽'
KG   = '#,##0.000'
PCT  = '0.0%'
NUM0 = '#,##0'
NUM2 = '#,##0.00'

def side():
    s = Side(style='thin', color='BFBFBF')
    return Border(left=s, right=s, top=s, bottom=s)

def fill(hex_):
    return PatternFill("solid", fgColor=hex_)

def write(ws, row, col, value, fmt=None, bold=False, color=None,
          bg=None, align='center', wrap=False, italic=False):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(
        bold=bold, italic=italic,
        color=color or "000000",
        name='Calibri', size=10
    )
    c.alignment = Alignment(
        horizontal=align, vertical='center', wrap_text=wrap
    )
    c.border = side()
    if fmt:
        c.number_format = fmt
    if bg:
        c.fill = fill(bg)
    return c

def hdr(ws, row, col, text, bg=C_DARK, color=C_WHITE, bold=True,
        wrap=True, align='center', span=None):
    if span:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=col + span - 1)
    c = ws.cell(row=row, column=col, value=text)
    c.font      = Font(bold=bold, color=color, name='Calibri', size=10)
    c.alignment = Alignment(horizontal=align, vertical='center', wrap_text=wrap)
    c.border    = side()
    c.fill      = fill(bg)
    return c

def row_bg(r):
    return C_ALT if r % 2 == 0 else C_WHITE

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

# ── WORKBOOK ───────────────────────────────────────────────────────────────

wb = openpyxl.Workbook()
wb.remove(wb.active)

# ══════════════════════════════════════════════════════════════════════════
# ЛИСТ 1 — ТОВАРЫ
# ══════════════════════════════════════════════════════════════════════════

ws = wb.create_sheet("Товары")
ws.sheet_view.showGridLines = False
ws.freeze_panes = "A3"

# ── Шапка ─────────────────────────────────────────────────────────────────
# Строка 1: группы
groups = [
    (1,  3,  "ИДЕНТИФИКАЦИЯ",             C_DARK),
    (4,  5,  "СЕБЕСТОИМОСТЬ",             C_BLUE),
    (6,  6,  "РУЧКА",                     C_ORANGE),
    (7,  8,  "ВЕС",                       "4472C4"),
    (9,  10, "ДОСТАВКА",                  "70AD47"),
    (11, 11, "ПОЛНАЯ СЕБ-ТЬ / ШТ",       "C00000"),
    (12, 12, "ЦЕНА ПРОДАЖИ",             C_ORANGE),
    (13, 14, "ПРИБЫЛЬ",                   "375623"),
]
for c1, c2, title, bg in groups:
    hdr(ws, 1, c1, title, bg=bg, span=c2 - c1 + 1)

# Строка 2: колонки
COLS = [
    "Диаметр\n(мм)",
    "Длина\n(м)",
    "Кол-во\n(шт)",
    "Себ-ть/м\n(₽)",
    "Себ-ть троса\n1 шт (₽)",
    "Ручка\n(₽)",
    "Вес/м\n(кг)",
    "Вес позиции\n(кг)",
    "Доля доставки\nна позицию (₽)",
    "Доставка\n1 шт (₽)",
    "Полная себ-ть\n1 шт (₽)",
    "Цена продажи\n1 шт (₽)",
    "Прибыль\n1 шт (₽)",
    "Прибыль\nпозиции (₽)",
]
for i, h in enumerate(COLS, 1):
    hdr(ws, 2, i, h, bg="2E4057")

ws.row_dimensions[1].height = 22
ws.row_dimensions[2].height = 44
set_col_widths(ws, [10, 8, 9, 12, 16, 10, 9, 16, 20, 14, 19, 18, 14, 18])

# ── Данные ────────────────────────────────────────────────────────────────
for r_idx, p in enumerate(rows, 3):
    bg = row_bg(r_idx)
    data = [
        (p["diam"],        NUM0),
        (p["length"],      NUM0),
        (p["qty"],         NUM0),
        (p["cpm"],         RUB),
        (p["rope_cost"],   RUB),
        (p["handle"],      RUB),
        (p["wpm"],         KG),
        (p["weight_pos"],  KG),
        (p["deliv_pos"],   RUB),
        (p["deliv_per_unit"], RUB),
        (p["full_cost"],   RUB),
        (p["price"],       RUB),
        (p["profit_unit"], RUB),
        (p["profit_pos"],  RUB),
    ]
    for col, (val, fmt) in enumerate(data, 1):
        write(ws, r_idx, col, round(val, 2), fmt=fmt, bg=bg)
    ws.row_dimensions[r_idx].height = 18

# ── Итоговая строка ───────────────────────────────────────────────────────
tot = len(rows) + 3
hdr(ws, tot, 1, "ИТОГО", bg=C_DARK, span=3)
totals = {
    4:  (None,                                               RUB),
    5:  (total_rope_cost,                                    RUB),
    6:  (handle_total,                                       RUB),
    7:  (None,                                               None),
    8:  (total_weight,                                       KG),
    9:  (DELIVERY,                                           RUB),
    10: (None,                                               None),
    11: (total_batch_cost,                                   RUB),
    12: (revenue,                                            RUB),
    13: (None,                                               None),
    14: (sum(r["profit_pos"] for r in rows),                 RUB),
}
for col, (val, fmt) in totals.items():
    if val is not None:
        write(ws, tot, col, round(val, 2), fmt=fmt, bold=True,
              bg=C_DARK, color=C_WHITE)
    else:
        hdr(ws, tot, col, "—", bg=C_DARK)
ws.row_dimensions[tot].height = 22

# ══════════════════════════════════════════════════════════════════════════
# ЛИСТ 2 — СВОДКА
# ══════════════════════════════════════════════════════════════════════════

ws2 = wb.create_sheet("Сводка")
ws2.sheet_view.showGridLines = False
ws2.column_dimensions["A"].width = 42
ws2.column_dimensions["B"].width = 22
ws2.column_dimensions["C"].width = 28

def s2_title(row, text, bg=C_DARK):
    ws2.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    c = ws2.cell(row=row, column=1, value=text)
    c.font      = Font(bold=True, color=C_WHITE, name='Calibri', size=11)
    c.fill      = fill(bg)
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border    = side()
    ws2.row_dimensions[row].height = 24

def s2_row(row, label, value, fmt=RUB, bold=False,
           val_bg=None, lbl_bg=None, note=None):
    bg = row_bg(row)
    c = ws2.cell(row=row, column=1, value=label)
    c.font      = Font(bold=bold, name='Calibri', size=10)
    c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    c.border    = side()
    c.fill      = fill(lbl_bg or bg)

    v = ws2.cell(row=row, column=2, value=round(value, 2) if isinstance(value, float) else value)
    v.number_format = fmt
    v.font      = Font(bold=bold, name='Calibri', size=10)
    v.alignment = Alignment(horizontal='right', vertical='center')
    v.border    = side()
    v.fill      = fill(val_bg or lbl_bg or bg)

    if note:
        n = ws2.cell(row=row, column=3, value=note)
        n.font      = Font(italic=True, size=9, color="595959", name='Calibri')
        n.alignment = Alignment(horizontal='left', vertical='center', indent=1, wrap_text=True)
        n.border    = side()
        n.fill      = fill(bg)
    else:
        ws2.cell(row=row, column=3, value="").border = side()

    ws2.row_dimensions[row].height = 20

# Заголовок
ws2.merge_cells("A1:C1")
c = ws2.cell(row=1, column=1, value="ЮНИТ-ЭКОНОМИКА — ТРОС NN")
c.font = Font(bold=True, size=14, color=C_WHITE, name='Calibri')
c.fill = fill(C_DARK)
c.alignment = Alignment(horizontal='center', vertical='center')
ws2.row_dimensions[1].height = 34

# Шапка столбцов
hdr(ws2, 2, 1, "Показатель", bg="2E4057", align='left')
hdr(ws2, 2, 2, "Значение",   bg="2E4057")
hdr(ws2, 2, 3, "Примечание", bg="2E4057", align='left')
ws2.row_dimensions[2].height = 22

r = 3

# ─ Партия ─
s2_title(r, "ПАРТИЯ"); r += 1
s2_row(r, "Позиций (видов товара)",   len(rows),    fmt=NUM0); r += 1
s2_row(r, "Итого штук",               total_qty,    fmt=NUM0); r += 1
s2_row(r, "Общий вес партии",         total_weight, fmt=f'{KG} "кг"'); r += 1

# ─ Себестоимость ─
s2_title(r, "СЕБЕСТОИМОСТЬ ПАРТИИ", bg=C_BLUE); r += 1
s2_row(r, "Себ-ть тросов (всего)",
       total_rope_cost, note="цена/м × длина × кол-во"); r += 1
s2_row(r, "Ручки (всего)",
       handle_total, note=f"{HANDLE} ₽ × {total_qty} шт"); r += 1
s2_row(r, "Доставка",
       DELIVERY, note="распределена пропорционально весу"); r += 1
s2_row(r, "Итого себ-ть партии",
       total_batch_cost, bold=True, lbl_bg=C_HEADER, val_bg=C_HEADER,
       note="трос + ручки + доставка"); r += 1
s2_row(r, "Маркетинг",
       MARKETING, note="фиксированный расход"); r += 1
s2_row(r, "Итого все расходы",
       total_batch_cost + MARKETING, bold=True, lbl_bg="DAEEF3",
       val_bg="DAEEF3"); r += 1

# ─ Выручка ─
s2_title(r, "ВЫРУЧКА", bg="70AD47"); r += 1
s2_row(r, "Выручка (все позиции, наценка 40%)",
       revenue, bold=True, note="цена = себ-ть × 1.4"); r += 1

# ─ Налог и прибыль ─
s2_title(r, "НАЛОГ И ПРИБЫЛЬ", bg=C_ORANGE); r += 1
s2_row(r, "Налог УСН 6%",
       tax, note=f"{TAX_RATE*100:.0f}% от выручки"); r += 1
s2_row(r, "Выручка за вычетом налога",
       revenue - tax); r += 1

# ─ Главный блок ─
r += 1
ws2.merge_cells(f"A{r}:C{r}")
c = ws2.cell(row=r, column=1,
             value="👉  ПРИБЫЛЬ С ПАРТИИ ПОСЛЕ ВСЕХ РАСХОДОВ")
c.font = Font(bold=True, size=12, color=C_WHITE, name='Calibri')
c.fill = fill("C00000")
c.alignment = Alignment(horizontal='center', vertical='center')
c.border = side()
ws2.row_dimensions[r].height = 30; r += 1

s2_row(r, "Выручка",                 revenue,      lbl_bg=C_YELLOW, val_bg=C_YELLOW); r += 1
s2_row(r, "− Себестоимость партии",  total_batch_cost, lbl_bg=C_YELLOW, val_bg=C_YELLOW); r += 1
s2_row(r, "− Маркетинг",             MARKETING,    lbl_bg=C_YELLOW, val_bg=C_YELLOW); r += 1
s2_row(r, "− Налог УСН 6%",         tax,           lbl_bg=C_YELLOW, val_bg=C_YELLOW); r += 1

# Итог — чистая прибыль
ws2.merge_cells(f"A{r}:A{r}")
c = ws2.cell(row=r, column=1, value="= ЧИСТАЯ ПРИБЫЛЬ С ПАРТИИ")
c.font = Font(bold=True, size=12, color=C_WHITE, name='Calibri')
c.fill = fill("375623" if net_profit >= 0 else "C00000")
c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
c.border = side()

v = ws2.cell(row=r, column=2, value=round(net_profit, 2))
v.number_format = RUB
v.font = Font(bold=True, size=13, color=C_WHITE, name='Calibri')
v.fill = fill("375623" if net_profit >= 0 else "C00000")
v.alignment = Alignment(horizontal='right', vertical='center')
v.border = side()

note_text = f"Маржа: {margin_pct:.1f}%"
n = ws2.cell(row=r, column=3, value=note_text)
n.font = Font(bold=True, size=11, color=C_WHITE, name='Calibri')
n.fill = fill("375623" if net_profit >= 0 else "C00000")
n.alignment = Alignment(horizontal='center', vertical='center')
n.border = side()
ws2.row_dimensions[r].height = 30; r += 1

# ─ Аналитика ─
r += 1
s2_title(r, "АНАЛИТИКА ПОЗИЦИЙ", bg="4472C4"); r += 1

# Топ-3
ws2.merge_cells(f"A{r}:C{r}")
c = ws2.cell(row=r, column=1,
             value="ТОП-3 прибыльных → увеличить закупку")
c.font = Font(bold=True, color="375623", name='Calibri', size=10)
c.fill = fill(C_GFILL)
c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
c.border = side()
ws2.row_dimensions[r].height = 20; r += 1

for p in sorted_by_profit[:3]:
    ws2.merge_cells(f"A{r}:C{r}")
    text = (f"  ✔  {p['diam']} мм × {p['length']} м  │  "
            f"Себ-ть: {p['full_cost']:,.0f} ₽  │  "
            f"Цена: {p['price']:,.0f} ₽  │  "
            f"Прибыль с позиции: {p['profit_pos']:,.0f} ₽  "
            f"({p['qty']} шт)")
    c = ws2.cell(row=r, column=1, value=text)
    c.font = Font(color="375623", name='Calibri', size=10)
    c.fill = fill(C_GFILL)
    c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    c.border = side()
    ws2.row_dimensions[r].height = 18; r += 1

# Слабые 3
ws2.merge_cells(f"A{r}:C{r}")
c = ws2.cell(row=r, column=1,
             value="СЛАБЫЕ позиции → сократить или пересмотреть цену")
c.font = Font(bold=True, color=C_RED, name='Calibri', size=10)
c.fill = fill(C_RFILL)
c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
c.border = side()
ws2.row_dimensions[r].height = 20; r += 1

for p in sorted_by_profit[-3:]:
    ws2.merge_cells(f"A{r}:C{r}")
    text = (f"  ✖  {p['diam']} мм × {p['length']} м  │  "
            f"Себ-ть: {p['full_cost']:,.0f} ₽  │  "
            f"Цена: {p['price']:,.0f} ₽  │  "
            f"Прибыль с позиции: {p['profit_pos']:,.0f} ₽  "
            f"({p['qty']} шт)")
    c = ws2.cell(row=r, column=1, value=text)
    c.font = Font(color=C_RED, name='Calibri', size=10)
    c.fill = fill(C_RFILL)
    c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    c.border = side()
    ws2.row_dimensions[r].height = 18; r += 1

# Выводы
r += 1
s2_title(r, "ВЫВОДЫ", bg="2E4057"); r += 1

conclusions = [
    f"Наценка 40% даёт чистую прибыль с партии: {net_profit:,.0f} ₽  (маржа {margin_pct:.1f}%)",
    f"Самая доходная позиция: {sorted_by_profit[0]['diam']} мм × {sorted_by_profit[0]['length']} м — прибыль {sorted_by_profit[0]['profit_pos']:,.0f} ₽",
    f"Слабейшая позиция: {sorted_by_profit[-1]['diam']} мм × {sorted_by_profit[-1]['length']} м — прибыль {sorted_by_profit[-1]['profit_pos']:,.0f} ₽",
    f"Точка безубыточности: нужно выручить минимум {total_batch_cost + MARKETING:,.0f} ₽ (до налога)",
    f"Маркетинг {MARKETING:,} ₽ = {MARKETING/net_profit*100:.1f}% от чистой прибыли — окупается",
]
for line in conclusions:
    ws2.merge_cells(f"A{r}:C{r}")
    bg = row_bg(r)
    c = ws2.cell(row=r, column=1, value=f"  • {line}")
    c.font = Font(name='Calibri', size=10)
    c.fill = fill(bg)
    c.alignment = Alignment(horizontal='left', vertical='center',
                             indent=1, wrap_text=True)
    c.border = side()
    ws2.row_dimensions[r].height = 20; r += 1

# ── Порядок листов ─────────────────────────────────────────────────────────
wb._sheets = [ws2, ws]

OUT = "/Users/alexeytrushin/Документы программ/Projects/tros-nn/unit_economics.xlsx"
wb.save(OUT)

# ── Консоль: сверка цифр ───────────────────────────────────────────────────
print("=" * 60)
print(f"  {'Позиция':<20} {'Себ-ть/шт':>10} {'Цена':>10} {'Приб/шт':>10} {'Приб.поз':>10}")
print("-" * 60)
for p in rows:
    name = f"{p['diam']}мм × {p['length']}м"
    print(f"  {name:<20} {p['full_cost']:>10,.0f} {p['price']:>10,.0f} "
          f"{p['profit_unit']:>10,.0f} {p['profit_pos']:>10,.0f}")
print("=" * 60)
print(f"  Кол-во штук:        {total_qty}")
print(f"  Общий вес:          {total_weight:.2f} кг")
print(f"  Себ-ть партии:      {total_batch_cost:,.0f} ₽")
print(f"  Выручка:            {revenue:,.0f} ₽")
print(f"  Налог (6%):         {tax:,.0f} ₽")
print(f"  Маркетинг:          {MARKETING:,} ₽")
print(f"  ЧИСТАЯ ПРИБЫЛЬ:     {net_profit:,.0f} ₽  (маржа {margin_pct:.1f}%)")
print(f"\nФайл: {OUT}")
