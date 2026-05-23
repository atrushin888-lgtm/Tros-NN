"""
Юнит-экономика Трос NN — v2
Цель: ROI = 40% от вложенных средств (товар + доставка + маркетинг) после налога.

Логика расчёта цены:
  полная_себестоимость/шт = себестоимость_троса + доля_доставки + доля_маркетинга
  цена = полная_себестоимость × 1.40 / (1 − 0.06)
  => чистая_прибыль/шт = цена × 0.94 − полная_себестоимость = полная_себестоимость × 0.40
  => ROI = 40% от каждого вложенного рубля ✓
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── ИСХОДНЫЕ ДАННЫЕ ────────────────────────────────────────────────────────

CPM = {6: 370, 8: 385, 10: 450, 12: 505, 14: 585, 16: 650, 18: 997}

ROI_TARGET = 0.40          # желаемый ROI на вложенные средства
DELIVERY   = 5_000
MARKETING  = 10_000
TAX_RATE   = 0.06

# Мультипликатор цены: чтобы после налога получить ROI_TARGET на полную себестоимость
# price × (1 − tax) = full_cost × (1 + ROI)  →  price = full_cost × (1+ROI)/(1−tax)
PRICE_MULT = (1 + ROI_TARGET) / (1 - TAX_RATE)   # = 1.40 / 0.94 ≈ 1.4894

# (диаметр, длина, основная=True/доп=False)
ITEMS = [
    (6,  5,  True),
    (6,  10, True),
    (8,  5,  True),
    (8,  10, True),
    (10, 5,  True),
    (10, 15, True),
    (12, 10, True),
    (12, 15, True),
    (12, 20, True),
    (14, 15, True),
    (14, 20, True),
    (16, 10, True),
    (16, 15, True),
    (18, 5,  False),
    (18, 10, False),
    (18, 15, False),
]

# ── РАСЧЁТ ─────────────────────────────────────────────────────────────────

# Шаг 1: себестоимость тросов
_rope_rows = []
for diam, length, main in ITEMS:
    _rope_rows.append({'diam': diam, 'length': length, 'main': main,
                       'cpm': CPM[diam], 'rope_cost': CPM[diam] * length, 'qty': 1})

main_rope_total = sum(r['rope_cost'] for r in _rope_rows if r['main'])

# Шаг 2: распределяем доставку и маркетинг пропорционально себестоимости
rows = []
for r in _rope_rows:
    if r['main']:
        deliv_share = DELIVERY  * r['rope_cost'] / main_rope_total
        mktg_share  = MARKETING * r['rope_cost'] / main_rope_total
    else:
        # для доп. позиций — только для анализа, без аллокации
        deliv_share = 0
        mktg_share  = 0

    full_cost = r['rope_cost'] + deliv_share + mktg_share
    price     = full_cost * PRICE_MULT
    profit_after_tax = price * (1 - TAX_RATE) - full_cost   # = full_cost × 0.40
    roi_check = profit_after_tax / full_cost                  # должно быть ровно 0.40

    rows.append({
        'diam': r['diam'], 'length': r['length'], 'main': r['main'],
        'cpm': r['cpm'], 'qty': r['qty'],
        'rope_cost':   r['rope_cost'],
        'deliv_share': deliv_share,
        'mktg_share':  mktg_share,
        'full_cost':   full_cost,
        'price':       price,
        'profit':      profit_after_tax,
        'roi':         roi_check,
    })

main_rows = [r for r in rows if r['main']]
add_rows  = [r for r in rows if not r['main']]

total_investment = sum(r['full_cost'] for r in main_rows)   # товар+доставка+маркетинг
total_revenue    = sum(r['price']     for r in main_rows)
tax              = total_revenue * TAX_RATE
net_profit       = total_revenue - tax - total_investment
roi_actual       = net_profit / total_investment             # должно быть 0.40
net_margin       = net_profit / total_revenue

# ── СТИЛИ ──────────────────────────────────────────────────────────────────

DARK   = "1F2D3D"
BLUE   = "2E75B6"
ORANGE = "F26522"
GREEN  = "375623"
RED    = "9C0006"
LGRAY  = "F5F7FA"
LBLUE  = "DBEAFE"
LGREEN = "DCFCE7"
LRED   = "FEE2E2"
YELLOW = "FFFBEB"
WHITE  = "FFFFFF"
ALT    = "F0F4F8"

def bdr():
    s = Side(style='thin', color='D1D5DB')
    return Border(left=s, right=s, top=s, bottom=s)

def fill(h):
    return PatternFill("solid", fgColor=h)

def cell(ws, row, col, val, fmt=None, bold=False, italic=False,
         color="000000", bg=None, align='center', wrap=False, size=10):
    c = ws.cell(row=row, column=col, value=val)
    c.font      = Font(bold=bold, italic=italic, color=color,
                       name='Calibri', size=size)
    c.alignment = Alignment(horizontal=align, vertical='center',
                             wrap_text=wrap)
    c.border    = bdr()
    if fmt: c.number_format = fmt
    if bg:  c.fill = fill(bg)
    return c

def hdr(ws, row, col, text, bg=DARK, fg=WHITE, span=1, bold=True,
        align='center', size=10, italic=False, wrap=True):
    if span > 1:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=col+span-1)
    c = ws.cell(row=row, column=col, value=text)
    c.font      = Font(bold=bold, italic=italic, color=fg,
                       name='Calibri', size=size)
    c.fill      = fill(bg)
    c.alignment = Alignment(horizontal=align, vertical='center',
                             wrap_text=wrap)
    c.border    = bdr()
    return c

RUB  = '#,##0 ₽'
PCT  = '0.0%'
N0   = '#,##0'

def bg_row(r):
    return ALT if r % 2 == 0 else WHITE

# ── WORKBOOK ───────────────────────────────────────────────────────────────

wb = openpyxl.Workbook()
wb.remove(wb.active)

# ══════════════════════════════════════════════════════════════════════════
# ЛИСТ 1 — СВОДКА
# ══════════════════════════════════════════════════════════════════════════

ws2 = wb.create_sheet("Сводка")
ws2.sheet_view.showGridLines = False
ws2.column_dimensions['A'].width = 40
ws2.column_dimensions['B'].width = 22
ws2.column_dimensions['C'].width = 26

# Заголовок
ws2.merge_cells("A1:C1")
c = ws2.cell(row=1, column=1, value="ЮНИТ-ЭКОНОМИКА — ТРОС NN  (наценка 40%)")
c.font      = Font(bold=True, size=16, color=WHITE, name='Calibri')
c.fill      = fill(DARK)
c.alignment = Alignment(horizontal='center', vertical='center')
c.border    = bdr()
ws2.row_dimensions[1].height = 36

# Шапка
hdr(ws2, 2, 1, "Показатель", bg="2E4057", align='left')
hdr(ws2, 2, 2, "Значение",   bg="2E4057")
hdr(ws2, 2, 3, "Примечание", bg="2E4057", align='left')
ws2.row_dimensions[2].height = 22

r = 3

def s2_title(row, text, bg=DARK):
    ws2.merge_cells(f"A{row}:C{row}")
    c = ws2.cell(row=row, column=1, value=text)
    c.font      = Font(bold=True, size=11, color=WHITE, name='Calibri')
    c.fill      = fill(bg)
    c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    c.border    = bdr()
    ws2.row_dimensions[row].height = 24

def s2_row(row, label, val, fmt=RUB, note='', bold=False,
           lbg=None, vbg=None, nbg=None):
    bg = bg_row(row)
    c = ws2.cell(row=row, column=1, value=label)
    c.font      = Font(bold=bold, name='Calibri', size=10)
    c.alignment = Alignment(horizontal='left', vertical='center', indent=2)
    c.border    = bdr()
    c.fill      = fill(lbg or bg)

    v = ws2.cell(row=row, column=2,
                 value=round(val, 2) if isinstance(val, float) else val)
    v.number_format = fmt
    v.font      = Font(bold=bold, name='Calibri', size=10)
    v.alignment = Alignment(horizontal='right', vertical='center')
    v.border    = bdr()
    v.fill      = fill(vbg or lbg or bg)

    n = ws2.cell(row=row, column=3, value=note)
    n.font      = Font(italic=True, size=9, color="6B7280", name='Calibri')
    n.alignment = Alignment(horizontal='left', vertical='center',
                             indent=2, wrap_text=True)
    n.border    = bdr()
    n.fill      = fill(nbg or lbg or bg)
    ws2.row_dimensions[row].height = 20

# ─ Блок 1: Логика расчёта ─
s2_title(r, "ЛОГИКА РАСЧЁТА ЦЕН", bg=BLUE); r+=1
s2_row(r, "Целевой ROI (на вложенные средства)",  ROI_TARGET,   fmt=PCT,
       note="Сколько зарабатываем с каждого вложенного рубля — после всех расходов"); r+=1
s2_row(r, "Налог УСН",                             TAX_RATE,     fmt=PCT,
       note="Вычитается из выручки"); r+=1
s2_row(r, "Мультипликатор цены",                   PRICE_MULT,   fmt='0.0000',
       note=f"= (1 + {ROI_TARGET}) / (1 − {TAX_RATE}) = {PRICE_MULT:.4f}"); r+=1
s2_row(r, "Наценка к полной себестоимости",        PRICE_MULT-1, fmt=PCT,
       note="Итоговая наценка включает компенсацию налога"); r+=1

# ─ Блок 2: Вложения ─
total_rope   = sum(r2['rope_cost'] for r2 in main_rows)
s2_title(r, "ВЛОЖЕНИЯ (кол-во = 1 шт каждой позиции)", bg=BLUE); r+=1
s2_row(r, "Закупка тросов",  total_rope,        note="Себестоимость всех позиций"); r+=1
s2_row(r, "Доставка",        DELIVERY,           note="Распределена пропорционально в цены"); r+=1
s2_row(r, "Маркетинг",       MARKETING,          note="Распределён пропорционально в цены"); r+=1
s2_row(r, "Итого вложений",  total_investment,
       bold=True, lbg=LBLUE, vbg=LBLUE, nbg=LBLUE,
       note="Товар + доставка + маркетинг"); r+=1

# ─ Блок 3: Выручка ─
s2_title(r, "ВЫРУЧКА", bg="375623"); r+=1
s2_row(r, "Общая выручка",   total_revenue, bold=True,
       note=f"Себестоимость × {PRICE_MULT:.4f}"); r+=1
s2_row(r, "Налог УСН 6%",    tax,
       note=f"6% от {total_revenue:,.0f} ₽"); r+=1
s2_row(r, "Чистая выручка",  total_revenue - tax,
       note="После уплаты налога"); r+=1

# ─ Главный блок ─
r+=1
ws2.merge_cells(f"A{r}:C{r}")
c = ws2.cell(row=r, column=1,
             value="👉  ЧИСТАЯ ПРИБЫЛЬ ПРИ ПРОДАЖЕ ВСЕЙ ПАРТИИ")
c.font      = Font(bold=True, size=12, color=WHITE, name='Calibri')
c.fill      = fill("15803D")
c.alignment = Alignment(horizontal='center', vertical='center')
c.border    = bdr()
ws2.row_dimensions[r].height = 30; r+=1

s2_row(r, "Выручка",           total_revenue,   lbg=YELLOW, vbg=YELLOW, nbg=YELLOW); r+=1
s2_row(r, "− Налог УСН 6%",    tax,             lbg=YELLOW, vbg=YELLOW, nbg=YELLOW); r+=1
s2_row(r, "− Вложения",        total_investment,lbg=YELLOW, vbg=YELLOW, nbg=YELLOW,
       note="Товар + доставка + маркетинг"); r+=1

# Итог
net_bg = "15803D"
c = ws2.cell(row=r, column=1, value="= ЧИСТАЯ ПРИБЫЛЬ")
c.font      = Font(bold=True, size=12, color=WHITE, name='Calibri')
c.fill      = fill(net_bg)
c.alignment = Alignment(horizontal='left', vertical='center', indent=2)
c.border    = bdr()

v = ws2.cell(row=r, column=2, value=round(net_profit, 2))
v.number_format = RUB
v.font      = Font(bold=True, size=14, color=WHITE, name='Calibri')
v.fill      = fill(net_bg)
v.alignment = Alignment(horizontal='right', vertical='center')
v.border    = bdr()

n = ws2.cell(row=r, column=3,
             value=f"ROI = {roi_actual*100:.1f}% от вложений  |  Маржа = {net_margin*100:.1f}% от выручки")
n.font      = Font(bold=True, size=11, color=WHITE, name='Calibri')
n.fill      = fill(net_bg)
n.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
n.border    = bdr()
ws2.row_dimensions[r].height = 34; r+=1

# ─ Аналитика ─
r+=1
s2_title(r, "АНАЛИТИКА — КАКИЕ ПОЗИЦИИ ВЫГОДНЕЕ", bg="4472C4"); r+=1

p14_20 = next(x for x in main_rows if x['diam']==14 and x['length']==20)
p8_10  = next(x for x in main_rows if x['diam']==8  and x['length']==10)
p12_20 = next(x for x in main_rows if x['diam']==12 and x['length']==20)
p18_15 = next(x for x in add_rows  if x['diam']==18 and x['length']==15)

analyses = [
    ("ROI на каждую позицию", f"Одинаков для всех — ровно {ROI_TARGET*100:.0f}% от вложенных средств. "
                               "Это гарантия: каждая позиция окупается с одинаковой отдачей."),
    ("По абс. прибыли / шт",  f"14мм×20м — {p14_20['profit']:,.0f} ₽/шт  |  "
                               f"12мм×20м — {p12_20['profit']:,.0f} ₽/шт  |  "
                               f"16мм×15м — {next(x for x in main_rows if x['diam']==16 and x['length']==15)['profit']:,.0f} ₽/шт"),
    ("Для масштабирования",    f"8мм×10м — самый ходовой, {p8_10['profit']:,.0f} ₽/шт. "
                               "Высокий объём продаж × хорошая прибыль = лучший выбор для роста."),
    ("18мм (доп. позиция)",    f"Прибыль {p18_15['profit']:,.0f} ₽/шт (15 м) — максимум в линейке. "
                               f"ROI тот же {ROI_TARGET*100:.0f}%, но нишевый спрос."),
    ("Вывод",                  "Масштабировать: 8мм и 10мм — массовый спрос, стабильный оборот. "
                               "Максимальный доход: 14мм×20м и 12мм×20м. "
                               "18мм добавить при наличии проф. клиентов."),
]

for label, text in analyses:
    bg = bg_row(r)
    ws2.merge_cells(f"B{r}:C{r}")
    c = ws2.cell(row=r, column=1, value=label)
    c.font      = Font(bold=True, name='Calibri', size=10)
    c.alignment = Alignment(horizontal='left', vertical='center', indent=2)
    c.border    = bdr()
    c.fill      = fill(bg)

    t = ws2.cell(row=r, column=2, value=text)
    t.font      = Font(name='Calibri', size=10)
    t.alignment = Alignment(horizontal='left', vertical='center',
                             indent=1, wrap_text=True)
    t.border    = bdr()
    t.fill      = fill(bg)
    ws2.row_dimensions[r].height = 32
    r+=1

# ══════════════════════════════════════════════════════════════════════════
# ЛИСТ 2 — ТОВАРЫ
# ══════════════════════════════════════════════════════════════════════════

ws1 = wb.create_sheet("Товары")
ws1.sheet_view.showGridLines = False
ws1.freeze_panes = "A3"

# Строка 1: группы
grp = [
    (1,  4,  "ПОЗИЦИЯ",                                      DARK),
    (5,  7,  "СЕБЕСТОИМОСТЬ (с долей доставки и маркетинга)", BLUE),
    (8,  8,  f"ЦЕНА (×{PRICE_MULT:.4f}, ROI 40% после налога)", ORANGE),
    (9,  10, "ПРИБЫЛЬ ПОСЛЕ НАЛОГА",                          "15803D"),
]
for c1, c2, t, bg in grp:
    hdr(ws1, 1, c1, t, bg=bg, span=c2-c1+1)

# Строка 2: колонки
COLH = [
    "Диам.\n(мм)", "Длина\n(м)", "Кол-во\n(шт)", "Себ-ть/м\n(₽)",
    "Себ-ть\nтроса (₽)", "Доля дост.+\nмаркет. (₽)", "Полная себ-ть\n1 шт (₽)",
    "Цена\nпродажи (₽)",
    "Прибыль\n1 шт (₽)", "Прибыль\n(с кол-ва)",
]
for i, h in enumerate(COLH, 1):
    hdr(ws1, 2, i, h, bg="2E4057")

ws1.row_dimensions[1].height = 22
ws1.row_dimensions[2].height = 44

WIDTHS = [10, 9, 11, 13, 14, 16, 16, 16, 15, 15]
for i, w in enumerate(WIDTHS, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w

# ── Основные позиции ──────────────────────────────────────────────────────
r_idx = 3
for p in main_rows:
    bg = bg_row(r_idx)
    fixed_share = p['deliv_share'] + p['mktg_share']
    vals = [
        (p['diam'],            None),
        (p['length'],          None),
        (p['qty'],             N0),
        (p['cpm'],             RUB),
        (p['rope_cost'],       RUB),
        (fixed_share,          RUB),
        (p['full_cost'],       RUB),
        (p['price'],           RUB),
        (p['profit'],          RUB),
        (p['profit']*p['qty'], RUB),
    ]
    for col, (val, fmt) in enumerate(vals, 1):
        cell(ws1, r_idx, col, round(val, 2) if isinstance(val, float) else val,
             fmt=fmt, bg=bg)
    ws1.row_dimensions[r_idx].height = 18
    r_idx += 1

# ── Итого основных ────────────────────────────────────────────────────────
hdr(ws1, r_idx, 1, "ИТОГО (основные позиции)", bg=DARK, span=4)
totals_main = {
    5:  sum(r['rope_cost'] * r['qty'] for r in main_rows),
    6:  DELIVERY + MARKETING,
    7:  total_investment,
    8:  total_revenue,
    9:  None,
    10: sum(r['profit'] * r['qty'] for r in main_rows),
}
for col, val in totals_main.items():
    if val is not None:
        cell(ws1, r_idx, col, round(val, 2), fmt=RUB, bold=True, bg=DARK, color=WHITE)
    else:
        hdr(ws1, r_idx, col, "—", bg=DARK)
ws1.row_dimensions[r_idx].height = 22
r_idx += 2

# ── 18 мм — дополнительный анализ ────────────────────────────────────────
ws1.merge_cells(f"A{r_idx}:{get_column_letter(len(COLH))}{r_idx}")
c = ws1.cell(row=r_idx, column=1,
             value="18 мм — ДОПОЛНИТЕЛЬНЫЙ АНАЛИЗ (без аллокации расходов — чисто по марже)")
c.font      = Font(bold=True, size=10, color=WHITE, name='Calibri')
c.fill      = fill("7C3AED")
c.alignment = Alignment(horizontal='left', vertical='center', indent=2)
c.border    = bdr()
ws1.row_dimensions[r_idx].height = 22
r_idx += 1

for p in add_rows:
    bg = "EDE9FE"
    fixed_share = p['deliv_share'] + p['mktg_share']
    # для доп. позиций цена — просто rope_cost × PRICE_MULT (без аллокации)
    price_add  = p['rope_cost'] * PRICE_MULT
    profit_add = price_add * (1 - TAX_RATE) - p['rope_cost']
    vals = [
        (p['diam'],          None),
        (p['length'],        None),
        (p['qty'],           N0),
        (p['cpm'],           RUB),
        (p['rope_cost'],     RUB),
        (0,                  RUB),
        (p['rope_cost'],     RUB),
        (price_add,          RUB),
        (profit_add,         RUB),
        (profit_add,         RUB),
    ]
    for col, (val, fmt) in enumerate(vals, 1):
        cell(ws1, r_idx, col,
             round(val, 2) if isinstance(val, float) else val,
             fmt=fmt, bg=bg)
    ws1.row_dimensions[r_idx].height = 18
    r_idx += 1

# ── Примечание ────────────────────────────────────────────────────────────
r_idx += 1
ws1.merge_cells(f"A{r_idx}:{get_column_letter(len(COLH))}{r_idx}")
c = ws1.cell(row=r_idx, column=1,
             value="  ℹ  Столбец «Кол-во» можно менять вручную — данный расчёт показан для qty = 1 шт каждой позиции")
c.font      = Font(italic=True, size=9, color="6B7280", name='Calibri')
c.fill      = fill(LGRAY)
c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
c.border    = bdr()
ws1.row_dimensions[r_idx].height = 18

# ── Порядок листов ────────────────────────────────────────────────────────
wb._sheets = [ws2, ws1]

# ── Сохранить ─────────────────────────────────────────────────────────────
OUT = "/Users/alexeytrushin/Документы программ/Projects/tros-nn/unit_economics_v2.xlsx"
wb.save(OUT)

# ── Консоль ───────────────────────────────────────────────────────────────
print("=" * 72)
print(f"  {'Позиция':<18} {'Себ-ть':>10} {'Полн.себ':>10} {'Цена':>11} {'Приб/шт':>10} {'ROI':>7}")
print("-" * 72)
for p in main_rows:
    name = f"{p['diam']}мм × {p['length']}м"
    print(f"  {name:<18} {p['rope_cost']:>10,.0f} {p['full_cost']:>10,.0f} "
          f"{p['price']:>11,.0f} {p['profit']:>10,.0f} {p['roi']*100:>6.1f}%")
print("-" * 72)
for p in add_rows:
    price_a  = p['rope_cost'] * PRICE_MULT
    profit_a = price_a * (1 - TAX_RATE) - p['rope_cost']
    name = f"{p['diam']}мм × {p['length']}м [доп]"
    print(f"  {name:<18} {p['rope_cost']:>10,.0f} {'(без аллок)':>10} "
          f"{price_a:>11,.0f} {profit_a:>10,.0f} {ROI_TARGET*100:>6.1f}%")
print("=" * 72)
print(f"\n  МУЛЬТИПЛИКАТОР ЦЕНЫ:  {PRICE_MULT:.6f}  (= 1.40 / 0.94)")
print(f"  ОСНОВНЫЕ (qty=1 каждой):")
print(f"  Вложения итого:  {total_investment:>12,.0f} ₽  (товар + доставка + маркетинг)")
print(f"  Выручка:         {total_revenue:>12,.0f} ₽")
print(f"  Налог 6%:        {tax:>12,.0f} ₽")
print(f"  ЧИСТАЯ ПРИБЫЛЬ:  {net_profit:>12,.0f} ₽")
print(f"  ROI:             {roi_actual*100:>11.1f}%  ← должно быть ровно 40%")
print(f"  Маржа:           {net_margin*100:>11.1f}%  (от выручки)")
print(f"\n  Файл: {OUT}")
