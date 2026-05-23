"""
Добавляет лист "Реальные рыночные цены" в unit_economics.xlsx.
Расчёт: рыночные цены продажи (не обратный расчёт от маржи).
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── ДАННЫЕ ─────────────────────────────────────────────────────────────────

COST_PER_M = {6: 180, 8: 250, 10: 272, 12: 322, 14: 470, 16: 539}
MARKET_PER_M = {6: 337, 8: 450, 10: 505, 12: 580, 14: 658, 16: 750}

PURCHASES = [
    (8,  5,  8),
    (8,  10, 6),
    (6,  5,  4),
    (6,  10, 2),
    (10, 5,  3),
    (10, 10, 3),
    (10, 15, 2),
    (12, 10, 3),
    (12, 15, 2),
    (12, 20, 2),
    (14, 15, 2),
    (14, 20, 1),
    (16, 10, 1),
    (16, 15, 1),
]

DELIVERY   = 5_000
MARKETING  = 10_000
WELDER     = 20_000
TAX_RATE   = 0.06

# ── РАСЧЁТ ПОЗИЦИЙ ─────────────────────────────────────────────────────────

rows = []
for diam, length, qty in PURCHASES:
    cost_unit  = COST_PER_M[diam] * length
    sell_unit  = MARKET_PER_M[diam] * length
    profit_unit = sell_unit - cost_unit
    margin_pct  = profit_unit / sell_unit if sell_unit else 0
    rows.append({
        "diam":        diam,
        "length":      length,
        "qty":         qty,
        "cpm":         COST_PER_M[diam],
        "mpm":         MARKET_PER_M[diam],
        "cost_unit":   cost_unit,
        "sell_unit":   sell_unit,
        "profit_unit": profit_unit,
        "margin_pct":  margin_pct,
        "rev_pos":     sell_unit * qty,
        "profit_pos":  profit_unit * qty,
    })

total_rope_cost = sum(r["cost_unit"] * r["qty"] for r in rows)
total_revenue   = sum(r["rev_pos"] for r in rows)
total_gross_profit = sum(r["profit_pos"] for r in rows)
fixed_costs     = DELIVERY + MARKETING + WELDER
tax             = total_revenue * TAX_RATE
net_profit      = total_revenue - total_rope_cost - fixed_costs - tax
net_margin_pct  = net_profit / total_revenue if total_revenue else 0

# ── СТИЛИ ──────────────────────────────────────────────────────────────────

C_DARK   = "1F3864"
C_BLUE   = "2E75B6"
C_GREEN  = "375623"
C_ORANGE = "C55A11"
C_RED    = "9C0006"
C_ALT    = "EBF3FB"
C_WHITE  = "FFFFFF"
C_YELLOW = "FFF2CC"
C_GFILL  = "E2EFDA"
C_RFILL  = "FCE4D6"

RUB  = '#,##0 ₽'
PCT  = '0.0%'
NUM0 = '#,##0'

def thin_side():
    s = Side(style='thin', color='BFBFBF')
    return Border(left=s, right=s, top=s, bottom=s)

def fill(hex_):
    return PatternFill("solid", fgColor=hex_)

def cell(ws, row, col, value, fmt=None, bold=False, color=None,
         bg=None, align='center', wrap=False):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(bold=bold, color=color or "000000", name='Calibri', size=10)
    c.alignment = Alignment(horizontal=align, vertical='center', wrap_text=wrap)
    c.border = thin_side()
    if fmt:
        c.number_format = fmt
    if bg:
        c.fill = fill(bg)
    return c

def header_cell(ws, row, col, text, bg=C_DARK, color=C_WHITE, span=None, wrap=True):
    if span:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=col + span - 1)
    c = ws.cell(row=row, column=col, value=text)
    c.font = Font(bold=True, color=color, name='Calibri', size=10)
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=wrap)
    c.border = thin_side()
    c.fill = fill(bg)
    return c

def row_bg(r):
    return C_ALT if r % 2 == 0 else C_WHITE

def wide_row(ws, row, col, text, bg, color=C_WHITE, size=11, bold=True, span=11):
    ws.merge_cells(start_row=row, start_column=col,
                   end_row=row, end_column=col + span - 1)
    c = ws.cell(row=row, column=col, value=text)
    c.font = Font(bold=bold, color=color, name='Calibri', size=size)
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = thin_side()
    c.fill = fill(bg)
    return c

# ── ОТКРЫТЬ КНИГУ И ДОБАВИТЬ ЛИСТ ─────────────────────────────────────────

PATH = "/Users/alexeytrushin/Документы программ/Projects/tros-nn/unit_economics.xlsx"
wb = openpyxl.load_workbook(PATH)

SHEET_NAME = "Реальные рыночные цены"
if SHEET_NAME in wb.sheetnames:
    del wb[SHEET_NAME]

ws = wb.create_sheet(SHEET_NAME)
ws.sheet_view.showGridLines = False
ws.freeze_panes = "A4"

COLS_W = [10, 8, 9, 14, 18, 14, 18, 14, 11, 18, 18]
for i, w in enumerate(COLS_W, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# ── СТРОКА 1: ЗАГОЛОВОК ────────────────────────────────────────────────────
ws.merge_cells("A1:K1")
c = ws.cell(row=1, column=1, value="АНАЛИЗ ПРОДАЖ ПО РЫНОЧНЫМ ЦЕНАМ — ТРОС NN")
c.font = Font(bold=True, size=14, color=C_WHITE, name='Calibri')
c.fill = fill(C_DARK)
c.alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[1].height = 36

# ── СТРОКА 2: ГРУППЫ КОЛОНОК ───────────────────────────────────────────────
groups = [
    (1, 3,  "ПОЗИЦИЯ",          "2E4057"),
    (4, 6,  "СЕБЕСТОИМОСТЬ",    C_BLUE),
    (7, 9,  "ПРОДАЖА",         "70AD47"),
    (10, 11, "ПРИБЫЛЬ",        C_ORANGE),
]
for c1, c2, title, bg in groups:
    header_cell(ws, 2, c1, title, bg=bg, span=c2 - c1 + 1)
ws.row_dimensions[2].height = 22

# ── СТРОКА 3: НАЗВАНИЯ КОЛОНОК ─────────────────────────────────────────────
COLS = [
    "Диаметр\n(мм)",
    "Длина\n(м)",
    "Кол-во\n(шт)",
    "Закупочная\nцена/м (₽)",
    "Себ-ть\n1 шт (₽)",
    "Рыночная\nцена/м (₽)",
    "Цена\n1 шт (₽)",
    "Прибыль\n1 шт (₽)",
    "Маржа\n(%)",
    "Выручка\nпозиции (₽)",
    "Прибыль\nпозиции (₽)",
]
for i, h in enumerate(COLS, 1):
    header_cell(ws, 3, i, h, bg="2E4057", wrap=True)
ws.row_dimensions[3].height = 44

# ── СТРОКИ ДАННЫХ ──────────────────────────────────────────────────────────
for r_idx, p in enumerate(rows, 4):
    bg = row_bg(r_idx)
    is_loss = p["profit_unit"] < 0
    profit_bg = C_RFILL if is_loss else bg

    cell(ws, r_idx, 1,  p["diam"],        NUM0, bg=bg)
    cell(ws, r_idx, 2,  p["length"],      NUM0, bg=bg)
    cell(ws, r_idx, 3,  p["qty"],         NUM0, bg=bg)
    cell(ws, r_idx, 4,  p["cpm"],         RUB,  bg=bg)
    cell(ws, r_idx, 5,  p["cost_unit"],   RUB,  bg=bg)
    cell(ws, r_idx, 6,  p["mpm"],         RUB,  bg=bg)
    cell(ws, r_idx, 7,  p["sell_unit"],   RUB,  bg=bg, bold=True)
    cell(ws, r_idx, 8,  p["profit_unit"], RUB,  bg=profit_bg,
         bold=is_loss, color="9C0006" if is_loss else None)
    cell(ws, r_idx, 9,  p["margin_pct"],  PCT,  bg=profit_bg,
         color="9C0006" if is_loss else None)
    cell(ws, r_idx, 10, p["rev_pos"],     RUB,  bg=bg)
    cell(ws, r_idx, 11, p["profit_pos"],  RUB,  bg=profit_bg, bold=True,
         color="9C0006" if is_loss else "375623")
    ws.row_dimensions[r_idx].height = 18

# ── ИТОГОВАЯ СТРОКА ТАБЛИЦЫ ────────────────────────────────────────────────
tot_row = 4 + len(rows)
ws.merge_cells(start_row=tot_row, start_column=1, end_row=tot_row, end_column=3)
c = ws.cell(row=tot_row, column=1, value="ИТОГО ПАРТИЯ")
c.font = Font(bold=True, color=C_WHITE, name='Calibri', size=10)
c.fill = fill(C_DARK); c.border = thin_side()
c.alignment = Alignment(horizontal='center', vertical='center')
for col, (val, fmt) in [
    (4, (None,              None)),
    (5, (total_rope_cost,   RUB)),
    (6, (None,              None)),
    (7, (total_revenue,     RUB)),
    (8, (None,              None)),
    (9, (None,              None)),
    (10, (total_revenue,    RUB)),
    (11, (total_gross_profit, RUB)),
]:
    if val is not None:
        c2 = ws.cell(row=tot_row, column=col, value=round(val, 2))
        c2.number_format = fmt
        c2.font = Font(bold=True, color=C_WHITE, name='Calibri', size=10)
        c2.fill = fill(C_DARK); c2.border = thin_side()
        c2.alignment = Alignment(horizontal='right', vertical='center')
    else:
        c2 = ws.cell(row=tot_row, column=col, value="—")
        c2.font = Font(bold=True, color=C_WHITE, name='Calibri', size=10)
        c2.fill = fill(C_DARK); c2.border = thin_side()
        c2.alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[tot_row].height = 22

# ── СВОДКА БАТЧА ───────────────────────────────────────────────────────────
r = tot_row + 2

def summary_title(row, text, bg):
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=11)
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(bold=True, color=C_WHITE, name='Calibri', size=11)
    c.fill = fill(bg)
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = thin_side()
    ws.row_dimensions[row].height = 26

def summary_row(row, label, value, fmt=RUB, bold=False, lbl_bg=None, val_bg=None, note=None):
    bg = row_bg(row)
    lb = lbl_bg or bg
    vb = val_bg or lb

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    c = ws.cell(row=row, column=1, value=label)
    c.font = Font(bold=bold, name='Calibri', size=10)
    c.fill = fill(lb)
    c.alignment = Alignment(horizontal='left', vertical='center', indent=2)
    c.border = thin_side()

    ws.merge_cells(start_row=row, start_column=8, end_row=row, end_column=9)
    v = ws.cell(row=row, column=8, value=round(value, 2) if isinstance(value, float) else value)
    v.number_format = fmt
    v.font = Font(bold=bold, name='Calibri', size=10)
    v.fill = fill(vb)
    v.alignment = Alignment(horizontal='right', vertical='center')
    v.border = thin_side()

    if note:
        ws.merge_cells(start_row=row, start_column=10, end_row=row, end_column=11)
        n = ws.cell(row=row, column=10, value=note)
        n.font = Font(italic=True, size=9, color="595959", name='Calibri')
        n.fill = fill(bg)
        n.alignment = Alignment(horizontal='left', vertical='center', indent=1, wrap_text=True)
        n.border = thin_side()
    else:
        ws.merge_cells(start_row=row, start_column=10, end_row=row, end_column=11)
        c2 = ws.cell(row=row, column=10, value="")
        c2.border = thin_side(); c2.fill = fill(bg)

    ws.row_dimensions[row].height = 20

# Раздел: структура расходов
summary_title(r, "СТРУКТУРА РАСХОДОВ И ДОХОДОВ", C_BLUE); r += 1
summary_row(r, "Закупка тросов (себестоимость)",
            total_rope_cost, note="цена/м × длина × кол-во"); r += 1
summary_row(r, "Доставка закупки",
            DELIVERY, note="фиксированный расход"); r += 1
summary_row(r, "Маркетинг",
            MARKETING, note="фиксированный расход"); r += 1
summary_row(r, "Доработка / сварщик",
            WELDER, note="сборка, ручки, финишная обработка"); r += 1
summary_row(r, "Итого постоянные расходы",
            fixed_costs, bold=True,
            lbl_bg="DAEEF3", val_bg="DAEEF3",
            note="доставка + маркетинг + доработка"); r += 1
summary_row(r, "Выручка по рыночным ценам",
            total_revenue, bold=True,
            lbl_bg="E2EFDA", val_bg="E2EFDA"); r += 1
summary_row(r, "Налог УСН 6% от выручки",
            tax, note=f"{TAX_RATE*100:.0f}% × {total_revenue:,.0f} ₽"); r += 1

# Главный блок: чистая прибыль
r += 1
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=11)
c = ws.cell(row=r, column=1,
            value="ЧИСТАЯ ПРИБЫЛЬ СО ВСЕЙ ПАРТИИ (после всех расходов и налога)")
c.font = Font(bold=True, size=12, color=C_WHITE, name='Calibri')
c.fill = fill("C00000")
c.alignment = Alignment(horizontal='center', vertical='center')
c.border = thin_side()
ws.row_dimensions[r].height = 30; r += 1

# Детальная строчка разбивки
for label, val in [
    ("Выручка",                total_revenue),
    ("− Себестоимость тросов", -total_rope_cost),
    ("− Постоянные расходы",   -fixed_costs),
    ("− Налог УСН 6%",         -tax),
]:
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    c = ws.cell(row=r, column=1, value=label)
    c.font = Font(name='Calibri', size=10)
    c.fill = fill(C_YELLOW); c.border = thin_side()
    c.alignment = Alignment(horizontal='left', vertical='center', indent=2)

    ws.merge_cells(start_row=r, start_column=8, end_row=r, end_column=11)
    v = ws.cell(row=r, column=8, value=round(abs(val), 2))
    v.number_format = RUB
    v.font = Font(name='Calibri', size=10)
    v.fill = fill(C_YELLOW); v.border = thin_side()
    v.alignment = Alignment(horizontal='right', vertical='center')
    ws.row_dimensions[r].height = 20; r += 1

# Итог чистой прибыли
is_loss = net_profit < 0
profit_color = "C00000" if is_loss else C_GREEN

ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
c = ws.cell(row=r, column=1, value="= ЧИСТАЯ ПРИБЫЛЬ")
c.font = Font(bold=True, size=13, color=C_WHITE, name='Calibri')
c.fill = fill(profit_color); c.border = thin_side()
c.alignment = Alignment(horizontal='left', vertical='center', indent=2)

ws.merge_cells(start_row=r, start_column=8, end_row=r, end_column=9)
v = ws.cell(row=r, column=8, value=round(net_profit, 2))
v.number_format = RUB
v.font = Font(bold=True, size=14, color=C_WHITE, name='Calibri')
v.fill = fill(profit_color); v.border = thin_side()
v.alignment = Alignment(horizontal='right', vertical='center')

ws.merge_cells(start_row=r, start_column=10, end_row=r, end_column=11)
n = ws.cell(row=r, column=10,
            value=f"Маржа: {net_margin_pct*100:.1f}% от выручки")
n.font = Font(bold=True, size=11, color=C_WHITE, name='Calibri')
n.fill = fill(profit_color); n.border = thin_side()
n.alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[r].height = 34; r += 1

# Вывод — выгодно ли?
r += 1
verdict = ("ВЫВОД: По рыночным ценам продавать ВЫГОДНО."
           if net_profit >= 10_000
           else "ВЫВОД: Чистая прибыль мала — рыночные цены почти не покрывают все расходы.")
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=11)
c = ws.cell(row=r, column=1, value=verdict)
verdict_bg = C_GFILL if net_profit >= 10_000 else C_RFILL
verdict_fg = C_GREEN if net_profit >= 10_000 else C_RED
c.font = Font(bold=True, size=11, color=verdict_fg, name='Calibri')
c.fill = fill(verdict_bg); c.border = thin_side()
c.alignment = Alignment(horizontal='left', vertical='center', indent=2)
ws.row_dimensions[r].height = 26; r += 1

# Дополнительные цифры
for line in [
    f"Выручка по рыночным ценам: {total_revenue:,.0f} ₽",
    f"Себестоимость тросов: {total_rope_cost:,.0f} ₽  |  Постоянные расходы: {fixed_costs:,.0f} ₽  |  Налог: {tax:,.0f} ₽",
    f"Валовая прибыль (до фиксированных расходов и налога): {total_gross_profit:,.0f} ₽",
    f"Чистая прибыль: {net_profit:,.0f} ₽  (реальная маржа {net_margin_pct*100:.1f}%)",
]:
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=11)
    bg = row_bg(r)
    c = ws.cell(row=r, column=1, value=f"  • {line}")
    c.font = Font(name='Calibri', size=10)
    c.fill = fill(bg); c.border = thin_side()
    c.alignment = Alignment(horizontal='left', vertical='center',
                             indent=1, wrap_text=True)
    ws.row_dimensions[r].height = 20; r += 1

# ── ПОРЯДОК ЛИСТОВ: новый первым ───────────────────────────────────────────
# Переставляем: новый лист в начало
wb._sheets.remove(ws)
wb._sheets.insert(0, ws)

wb.save(PATH)

# ── КОНСОЛЬ ────────────────────────────────────────────────────────────────
print("=" * 72)
print(f"  {'Позиция':<18} {'Закуп/шт':>9} {'Продажа/шт':>11} {'Приб/шт':>9} {'Маржа':>7} {'Приб.поз':>10}")
print("-" * 72)
for p in rows:
    name = f"{p['diam']}мм × {p['length']}м"
    print(f"  {name:<18} {p['cost_unit']:>9,.0f} {p['sell_unit']:>11,.0f} "
          f"{p['profit_unit']:>9,.0f} {p['margin_pct']*100:>6.1f}% {p['profit_pos']:>10,.0f}")
print("=" * 72)
print(f"  Выручка:                 {total_revenue:>10,.0f} ₽")
print(f"  Себестоимость тросов:    {total_rope_cost:>10,.0f} ₽")
print(f"  Постоянные расходы:      {fixed_costs:>10,.0f} ₽  (дост+маркет+доработка)")
print(f"  Налог УСН 6%:            {tax:>10,.0f} ₽")
print(f"  ЧИСТАЯ ПРИБЫЛЬ:          {net_profit:>10,.0f} ₽  (маржа {net_margin_pct*100:.1f}%)")
print(f"\nЛист добавлен в: {PATH}")
